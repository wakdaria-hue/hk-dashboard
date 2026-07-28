"""Excel export styled to match the existing HK Cost Report workbook.

Matched conventions (inspected from HK_Cost_Report_June_2026.xlsx):
- Header: Arial 10 bold white text on a dark (#2B2621) fill, thin
  light-beige (#D8D2C4) border, centered.
- Currency cells: literal "€" prefix, thousands separator, 2 decimals.
- Hour cells: 1 decimal.
- Title: Arial 14 bold. Subtitle/caveat notes: Arial 10 italic gray (#6B6459).
- Totals row: bold.
"""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="2B2621")
_THIN = Side(style="thin", color="D8D2C4")
HEADER_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
TITLE_FONT = Font(name="Arial", size=14, bold=True)
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="6B6459")
BODY_FONT = Font(name="Arial", size=10)
TOTAL_FONT = Font(name="Arial", size=10, bold=True)
CENTER = Alignment(horizontal="center", vertical="center")

CURRENCY_FMT = "€#,##0.00"
HOURS_FMT = "0.0"


def _col_width(col_name: str, series: pd.Series) -> float:
    # Plain str(v) on each Python value (not Series.astype(str)/.map) - pandas'
    # Arrow-backed string dtype (pandas 3.x) doesn't reliably stringify NaN/None
    # through .astype(str).map(len), which crashes on any column with blanks
    # (e.g. cost cells for a month with no payroll rate on file yet).
    values_len = max((len(str(v)) for v in series), default=0)
    return max(10, min(30, max(len(str(col_name)), values_len) + 4))


def _write_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    title: str,
    subtitle: str | None,
    currency_cols: tuple[str, ...],
    hours_cols: tuple[str, ...],
    total_row: bool,
):
    ws = wb.create_sheet(sheet_name[:31])
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    row = 2
    if subtitle:
        ws.cell(row=row, column=1, value=subtitle).font = SUBTITLE_FONT
        row += 1
    row += 1  # blank spacer

    header_row = row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = CENTER

    for r_offset, (_, record) in enumerate(df.iterrows()):
        r = header_row + 1 + r_offset
        for col_idx, col_name in enumerate(df.columns, start=1):
            value = record[col_name]
            if pd.isna(value):
                value = None
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.font = BODY_FONT
            if col_name in currency_cols:
                cell.number_format = CURRENCY_FMT
            elif col_name in hours_cols:
                cell.number_format = HOURS_FMT

    if total_row and not df.empty:
        total_r = header_row + len(df) + 1
        ws.cell(row=total_r, column=1, value="Total").font = TOTAL_FONT
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in currency_cols or col_name in hours_cols:
                col_letter = get_column_letter(col_idx)
                cell = ws.cell(row=total_r, column=col_idx)
                cell.value = f"=SUM({col_letter}{header_row + 1}:{col_letter}{header_row + len(df)})"
                cell.font = TOTAL_FONT
                cell.number_format = CURRENCY_FMT if col_name in currency_cols else HOURS_FMT

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _col_width(col_name, df[col_name])

    return ws


def export_workbook(sheets: dict[str, dict]) -> bytes:
    """sheets: {sheet_name: {df, title, subtitle?, currency_cols?, hours_cols?, total_row?}}"""
    wb = Workbook()
    wb.remove(wb.active)
    for name, spec in sheets.items():
        _write_sheet(
            wb,
            sheet_name=name,
            df=spec["df"],
            title=spec.get("title", name),
            subtitle=spec.get("subtitle"),
            currency_cols=tuple(spec.get("currency_cols", ())),
            hours_cols=tuple(spec.get("hours_cols", ())),
            total_row=spec.get("total_row", True),
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
